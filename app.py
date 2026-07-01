from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
import base64
import hmac
import json
import math
import re
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
import requests
import streamlit as st


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DEFAULT_WORKBOOK = DATA_DIR / "TFF Joure, export tarieven geldig tm 30-06-2026.xlsx"
VAN_HEUGTEN_WORKBOOK = DATA_DIR / "Tarieven Calculator 2026 Van Heugten Transport (Duitsland EXPORT) (1).xlsx"
VAN_HEUGTEN_FR_WORKBOOK = DATA_DIR / "Tarieven Calculator per 01-07-2026 Van Heugten Transport (Frankrijk EXPORT).xlsx"
DROST_WORKBOOK = DATA_DIR / "Export Drost offerte 2026-1 (1).xlsx"
DEFAULT_WORKBOOKS = [DEFAULT_WORKBOOK, VAN_HEUGTEN_WORKBOOK, VAN_HEUGTEN_FR_WORKBOOK, DROST_WORKBOOK]
SETTINGS_FILE = APP_DIR / "settings.json"
GITHUB_SETTINGS_PATH = "cloud_settings.json"
GITHUB_REPO = "wiersmakevin7-oss/tariefzoeker"
GITHUB_BRANCH = "main"
DEFAULT_SETTINGS = {"diesel_by_carrier": {}, "road_charge_by_carrier": {}, "margin_pct": 0.0}
SKIP_SHEETS = {"Voorblad", "Algemene Voorwaarden", "Toeslagen en condities", "Contacts"}


@dataclass(frozen=True)
class RateRow:
    country: str
    zone: str
    max_kg: float | None
    max_ldm: float | None
    max_eu_pallets: float | None
    max_block_pallets: float | None
    base_price: float
    transit: str | None
    carrier: str
    source: str


DEFAULT_COUNTRIES = [
    "AT",
    "BE",
    "BG",
    "CH",
    "CZ",
    "DE",
    "DK",
    "EE",
    "ES",
    "FI",
    "FR",
    "GB",
    "GR",
    "HR",
    "HU",
    "IT",
    "LT",
    "LV",
    "NO",
    "PL",
    "PT",
    "RO",
    "SE",
    "SI",
    "SK",
]


def as_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace("EUR", "").replace("€", "").replace(" ", "")
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        return float(text)
    except ValueError:
        return None


def clean_postcode(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def format_money(value: float) -> str:
    return f"€ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def format_decimal(value: float, decimals: int = 3) -> str:
    return f"{value:.{decimals}f}".replace(".", ",")


def format_zone(value: Any) -> str:
    number = as_number(value)
    if number is not None and number.is_integer():
        return str(int(number))
    return str(value).strip().upper()


def workbook_carrier(wb: openpyxl.Workbook, fallback: str) -> str:
    if "Voorblad" not in wb.sheetnames:
        return fallback
    sheet = wb["Voorblad"]
    for cell in ("E5", "A5"):
        value = sheet[cell].value
        if value:
            return str(value).strip()
    return fallback


def col_to_index(col: str) -> int:
    total = 0
    for char in col:
        total = total * 26 + (ord(char.upper()) - 64)
    return total


def split_cell_ref(ref: str) -> tuple[int, int]:
    match = re.match(r"([A-Z]+)([0-9]+)", ref.upper())
    if not match:
        return 0, 0
    return int(match.group(2)), col_to_index(match.group(1))


def read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings: list[str] = []
    for item in root.findall("x:si", ns):
        parts = [node.text or "" for node in item.findall(".//x:t", ns)]
        strings.append("".join(parts))
    return strings


def workbook_sheets(zf: zipfile.ZipFile) -> list[tuple[str, str]]:
    ns_main = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rels = {
        rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
        for rel in rels_root.findall("r:Relationship", ns_rel)
    }
    sheets: list[tuple[str, str]] = []
    for sheet in wb_root.findall(".//x:sheet", ns_main):
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rels.get(rel_id or "")
        if not target:
            continue
        path = target if target.startswith("xl/") else f"xl/{target}"
        sheets.append((sheet.attrib["name"], path))
    return sheets


def read_sheet_cells(
    zf: zipfile.ZipFile,
    path: str,
    shared_strings: list[str],
    max_row: int = 80,
) -> dict[tuple[int, int], Any]:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(zf.read(path))
    cells: dict[tuple[int, int], Any] = {}
    for row in root.findall(".//x:sheetData/x:row", ns):
        row_number = int(row.attrib.get("r", "0"))
        if row_number > max_row:
            continue
        for cell in row.findall("x:c", ns):
            ref = cell.attrib.get("r", "")
            row_idx, col_idx = split_cell_ref(ref)
            if row_idx == 0 or row_idx > max_row:
                continue
            cell_type = cell.attrib.get("t")
            value_node = cell.find("x:v", ns)
            if cell_type == "inlineStr":
                text = "".join(node.text or "" for node in cell.findall(".//x:t", ns))
                cells[(row_idx, col_idx)] = text
            elif value_node is not None:
                raw = value_node.text or ""
                if cell_type == "s":
                    idx = int(raw)
                    cells[(row_idx, col_idx)] = shared_strings[idx] if idx < len(shared_strings) else raw
                else:
                    cells[(row_idx, col_idx)] = as_number(raw) if as_number(raw) is not None else raw
    return cells


def fast_workbook_carrier(cells: dict[tuple[int, int], Any], fallback: str) -> str:
    for ref in ((5, 5), (5, 1)):
        value = cells.get(ref)
        if value:
            return str(value).strip()
    return fallback


def sheet_transits(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, str]:
    transits: dict[str, str] = {}
    for row in range(1, min(ws.max_row, 45) + 1):
        first = str(ws.cell(row=row, column=1).value or "").lower()
        if "groupage" in first:
            for col in range(5, ws.max_column + 1):
                header = ws.cell(row=6, column=col).value
                value = ws.cell(row=row, column=col).value
                if header is not None and value is not None:
                    transits[str(header).strip().upper()] = str(value).strip()
            break
    return transits


def parse_workbook(path: Path) -> list[RateRow]:
    if "Van Heugten" in path.name:
        return parse_van_heugten_workbook(path)
    if "Drost" in path.name:
        return parse_drost_workbook(path)

    rows: list[RateRow] = []
    with zipfile.ZipFile(path) as zf:
        shared_strings = read_shared_strings(zf)
        sheets = workbook_sheets(zf)
        carrier = path.stem
        sheet_cells: dict[str, dict[tuple[int, int], Any]] = {}
        for sheet_name, sheet_path in sheets:
            if sheet_name == "Voorblad":
                cells = read_sheet_cells(zf, sheet_path, shared_strings, max_row=25)
                carrier = fast_workbook_carrier(cells, path.stem)
                break

        for sheet_name, sheet_path in sheets:
            if sheet_name in SKIP_SHEETS:
                continue
            cells = sheet_cells.get(sheet_name) or read_sheet_cells(zf, sheet_path, shared_strings, max_row=80)
            if str(cells.get((6, 1), "")).strip().lower() != "kg":
                continue
            zones = [
                (col, format_zone(cells[(6, col)]))
                for col in range(5, 140)
                if cells.get((6, col)) not in (None, "")
            ]
            transits: dict[str, str] = {}
            for row_idx in range(1, 46):
                first = str(cells.get((row_idx, 1), "")).lower()
                if "groupage" in first:
                    for col, zone in zones:
                        value = cells.get((row_idx, col))
                        if value not in (None, ""):
                            transits[zone] = str(value).strip()
                    break

            for row_idx in range(7, 81):
                max_kg = as_number(cells.get((row_idx, 1)))
                max_ldm = as_number(cells.get((row_idx, 2)))
                max_eu = as_number(cells.get((row_idx, 3)))
                max_block = as_number(cells.get((row_idx, 4)))
                if max_kg is None and max_ldm is None:
                    continue

                for col, zone in zones:
                    price = as_number(cells.get((row_idx, col)))
                    if price is None:
                        continue
                    rows.append(
                        RateRow(
                            country=sheet_name.upper(),
                            zone=zone,
                            max_kg=max_kg,
                            max_ldm=max_ldm,
                            max_eu_pallets=max_eu,
                            max_block_pallets=max_block,
                            base_price=price,
                            transit=transits.get(zone),
                            carrier=carrier,
                            source=path.name,
                        )
                    )
    return rows


def parse_drost_workbook(path: Path) -> list[RateRow]:
    zone_prefixes = {
        5: ["6"],
        6: ["4", "5"],
        7: ["1", "2", "3"],
        8: ["7", "8"],
        9: ["9"],
    }
    rows: list[RateRow] = []
    with zipfile.ZipFile(path) as zf:
        shared_strings = read_shared_strings(zf)
        sheets = workbook_sheets(zf)
        if not sheets:
            return rows
        _, sheet_path = sheets[0]
        cells = read_sheet_cells(zf, sheet_path, shared_strings, max_row=80)
        if str(cells.get((6, 1), "")).strip().upper() != "LDM":
            return rows

        for row_idx in range(7, 81):
            max_ldm = as_number(cells.get((row_idx, 1)))
            max_eu = as_number(cells.get((row_idx, 2)))
            max_block = as_number(cells.get((row_idx, 3)))
            max_kg = as_number(cells.get((row_idx, 4)))
            if max_ldm is None and max_kg is None:
                continue

            for col, prefixes in zone_prefixes.items():
                price = as_number(cells.get((row_idx, col)))
                if price is None:
                    continue
                for prefix in prefixes:
                    rows.append(
                        RateRow(
                            country="AT",
                            zone=prefix,
                            max_kg=max_kg,
                            max_ldm=max_ldm,
                            max_eu_pallets=max_eu,
                            max_block_pallets=max_block,
                            base_price=price,
                            transit=None,
                            carrier="Drost",
                            source=path.name,
                        )
                    )
    return rows


def parse_van_heugten_workbook(path: Path) -> list[RateRow]:
    rows: list[RateRow] = []
    country_map = {"UK": "GB"}
    filename_country_filters = {
        "duitsland": {"DE"},
        "frankrijk": {"FR"},
    }
    allowed_countries: set[str] | None = None
    lower_name = path.name.lower()
    for marker, countries in filename_country_filters.items():
        if marker in lower_name:
            allowed_countries = countries
            break
    with zipfile.ZipFile(path) as zf:
        shared_strings = read_shared_strings(zf)
        sheets = workbook_sheets(zf)
        for sheet_name, sheet_path in sheets:
            if not sheet_name.endswith("-Exp") or sheet_name.endswith("(O)"):
                continue
            country_code = country_map.get(sheet_name.split("-", 1)[0].upper(), sheet_name.split("-", 1)[0].upper())
            if allowed_countries is not None and country_code not in allowed_countries:
                continue
            cells = read_sheet_cells(zf, sheet_path, shared_strings, max_row=110)
            if str(cells.get((5, 1), "")).strip().lower() != "postcode":
                continue

            zones = [
                (col, format_zone(cells[(5, col)]))
                for col in range(16, 150)
                if cells.get((5, col)) not in (None, "", 0)
            ]
            for row_idx in range(7, 111):
                max_ldm = as_number(cells.get((row_idx, 14)))
                max_kg = as_number(cells.get((row_idx, 15)))
                if max_ldm is None and max_kg is None:
                    continue
                for col, zone in zones:
                    price = as_number(cells.get((row_idx, col)))
                    if price is None or price <= 0:
                        continue
                    transit = cells.get((row_idx, 4))
                    rows.append(
                        RateRow(
                            country=country_code,
                            zone=zone,
                            max_kg=max_kg,
                            max_ldm=max_ldm,
                            max_eu_pallets=None,
                            max_block_pallets=None,
                            base_price=price,
                            transit=str(transit).strip() if transit not in (None, "") else None,
                            carrier="Van Heugten Transport",
                            source=path.name,
                        )
                    )
    return rows


@st.cache_data(show_spinner=False)
def load_default_rates(path_text: str) -> list[RateRow]:
    return parse_workbook(Path(path_text))


@st.cache_data(show_spinner=False)
def load_default_rate_books(paths: tuple[str, ...]) -> list[RateRow]:
    rows: list[RateRow] = []
    for path_text in paths:
        rows.extend(parse_workbook(Path(path_text)))
    return rows


def load_uploaded_rates(files: list[Any]) -> list[RateRow]:
    all_rows: list[RateRow] = []
    for file in files:
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file.getvalue())
            tmp_path = Path(tmp.name)
        try:
            all_rows.extend(parse_workbook(tmp_path))
        finally:
            tmp_path.unlink(missing_ok=True)
    return all_rows


def postcode_matches(zone: str, postcode: str, country: str) -> bool:
    if zone == country:
        return True
    clean_zone = clean_postcode(zone)
    return bool(clean_zone and postcode.startswith(clean_zone))


def best_zone(country_rows: list[RateRow], postcode: str, country: str) -> str | None:
    zones = sorted({row.zone for row in country_rows}, key=lambda z: len(clean_postcode(z)), reverse=True)
    for zone in zones:
        if postcode_matches(zone, postcode, country):
            return zone
    return zones[0] if len(zones) == 1 else None


def row_fits(row: RateRow, kg: float, ldm: float, eu_pallets: float, block_pallets: float) -> bool:
    checks = [
        row.max_kg is None or kg <= row.max_kg,
        row.max_ldm is None or ldm <= row.max_ldm,
    ]
    if eu_pallets > 0 and block_pallets == 0 and row.max_eu_pallets is None and row.max_block_pallets is not None:
        return False
    if block_pallets > 0 and eu_pallets == 0 and row.max_block_pallets is None and row.max_eu_pallets is not None:
        return False
    if eu_pallets > 0 and row.max_eu_pallets is not None:
        checks.append(eu_pallets <= row.max_eu_pallets)
    if block_pallets > 0 and row.max_block_pallets is not None:
        checks.append(block_pallets <= row.max_block_pallets)
    return all(checks)


def find_rates(
    rows: list[RateRow],
    country: str,
    postcode: str,
    kg: float,
    ldm: float,
    eu_pallets: float,
    block_pallets: float,
) -> tuple[str | None, list[RateRow]]:
    country = country.upper()
    postcode = clean_postcode(postcode)
    country_rows = [row for row in rows if row.country == country]
    matched_zones: dict[str, str] = {}
    best_by_carrier: dict[str, RateRow] = {}

    for carrier in sorted({row.carrier for row in country_rows}):
        carrier_rows = [row for row in country_rows if row.carrier == carrier]
        zone = best_zone(carrier_rows, postcode, country)
        if zone is None:
            continue
        carrier_matches = [
            row
            for row in carrier_rows
            if row.zone == zone and row_fits(row, kg, ldm, eu_pallets, block_pallets)
        ]
        if not carrier_matches:
            continue
        matched_zones[carrier] = zone
        best_by_carrier[carrier] = sorted(carrier_matches, key=lambda item: item.base_price)[0]

    if not best_by_carrier:
        return None, []
    zone_display = ", ".join(f"{carrier}: {zone}" for carrier, zone in matched_zones.items())
    return zone_display, sorted(best_by_carrier.values(), key=lambda item: item.base_price)


def auto_ldm(length_cm: float, width_cm: float, pieces: int) -> float:
    if length_cm <= 0 or width_cm <= 0 or pieces <= 0:
        return 0.0
    return pieces * (length_cm / 100) * (width_cm / 100) / 2.4


def pallet_kind_counts(pallet_rows: list[dict[str, Any]]) -> tuple[float, float]:
    eu_pallets = 0.0
    block_pallets = 0.0
    for row in pallet_rows:
        qty = as_number(row.get("Aantal")) or 0
        kind = str(row.get("Type") or "").lower()
        if "blok" in kind:
            block_pallets += qty
        elif "euro" in kind:
            eu_pallets += qty
    return eu_pallets, block_pallets


def shipment_totals(pallet_rows: list[dict[str, Any]]) -> dict[str, float]:
    totals = {"pallets": 0.0, "kg": 0.0, "ldm": 0.0, "m3": 0.0}
    for row in pallet_rows:
        qty = as_number(row.get("Aantal")) or 0
        length = as_number(row.get("Lengte cm")) or 0
        width = as_number(row.get("Breedte cm")) or 0
        height = as_number(row.get("Hoogte cm")) or 0
        weight_each = as_number(row.get("Kg per pallet")) or 0
        if qty <= 0:
            continue
        totals["pallets"] += qty
        totals["kg"] += qty * weight_each
        totals["ldm"] += auto_ldm(length, width, int(qty))
        totals["m3"] += qty * (length / 100) * (width / 100) * (height / 100)
    return totals


def total_with_diesel(base_price: float, diesel_pct: float) -> float:
    return base_price * (1 + diesel_pct / 100)


def road_tax_amount(base_price: float, road_tax_pct: float) -> float:
    return base_price * (road_tax_pct / 100)


def total_purchase(base_price: float, diesel_pct: float, road_tax_pct: float) -> float:
    return total_with_diesel(base_price, diesel_pct) + road_tax_amount(base_price, road_tax_pct)


def app_password() -> str:
    try:
        return str(st.secrets.get("APP_PASSWORD", "")).strip()
    except Exception:
        return ""


def require_login() -> None:
    password = app_password()
    if not password:
        return
    if st.session_state.get("authenticated"):
        return

    st.title("TFF tariefzoeker")
    st.caption("Log in om de tariefzoeker te gebruiken.")
    entered_password = st.text_input("Wachtwoord", type="password")
    if st.button("Inloggen", type="primary"):
        if hmac.compare_digest(entered_password, password):
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Wachtwoord klopt niet.")
    st.stop()


def normalize_settings(data: dict[str, Any] | None) -> dict[str, Any]:
    data = data or {}
    return {
        "diesel_by_carrier": data.get("diesel_by_carrier", {}),
        "road_charge_by_carrier": data.get("road_charge_by_carrier", {}),
        "margin_pct": as_number(data.get("margin_pct")) or 0.0,
    }


def github_token() -> str:
    try:
        return str(st.secrets.get("GITHUB_TOKEN", "")).strip()
    except Exception:
        return ""


def github_settings_url() -> str:
    return f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_SETTINGS_PATH}"


def github_headers(token: str) -> dict[str, str]:
    return {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def load_github_settings(token: str) -> dict[str, Any]:
    response = requests.get(
        github_settings_url(),
        headers=github_headers(token),
        params={"ref": GITHUB_BRANCH},
        timeout=15,
    )
    if response.status_code == 404:
        return normalize_settings(DEFAULT_SETTINGS)
    response.raise_for_status()
    payload = response.json()
    raw = base64.b64decode(payload["content"]).decode("utf-8")
    return normalize_settings(json.loads(raw))


def save_github_settings(
    token: str,
    diesel_by_carrier: dict[str, float],
    road_charge_by_carrier: dict[str, float],
    margin_pct: float,
) -> None:
    settings_payload = json.dumps(
        {
            "diesel_by_carrier": diesel_by_carrier,
            "road_charge_by_carrier": road_charge_by_carrier,
            "margin_pct": margin_pct,
        },
        indent=2,
        ensure_ascii=False,
    )
    current = requests.get(
        github_settings_url(),
        headers=github_headers(token),
        params={"ref": GITHUB_BRANCH},
        timeout=15,
    )
    body: dict[str, Any] = {
        "message": "Update tariefzoeker cloud settings",
        "content": base64.b64encode(settings_payload.encode("utf-8")).decode("ascii"),
        "branch": GITHUB_BRANCH,
    }
    if current.status_code == 200:
        body["sha"] = current.json()["sha"]
    elif current.status_code != 404:
        current.raise_for_status()
    response = requests.put(github_settings_url(), headers=github_headers(token), json=body, timeout=15)
    response.raise_for_status()


def load_settings() -> dict[str, Any]:
    token = github_token()
    if token:
        try:
            return load_github_settings(token)
        except (requests.RequestException, json.JSONDecodeError, KeyError, UnicodeDecodeError) as exc:
            st.warning(f"Cloud-instellingen konden niet worden geladen, lokale instellingen worden gebruikt: {exc}")
    if not SETTINGS_FILE.exists():
        return normalize_settings(DEFAULT_SETTINGS)
    try:
        with SETTINGS_FILE.open("r", encoding="utf-8") as file:
            data = json.load(file)
    except (OSError, json.JSONDecodeError):
        return normalize_settings(DEFAULT_SETTINGS)
    return normalize_settings(data)


def save_settings(
    diesel_by_carrier: dict[str, float],
    road_charge_by_carrier: dict[str, float],
    margin_pct: float,
) -> str:
    token = github_token()
    if token:
        save_github_settings(token, diesel_by_carrier, road_charge_by_carrier, margin_pct)
        return "Opgeslagen in GitHub cloud settings."
    SETTINGS_FILE.write_text(
        json.dumps(
            {
                "diesel_by_carrier": diesel_by_carrier,
                "road_charge_by_carrier": road_charge_by_carrier,
                "margin_pct": margin_pct,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return "Lokaal opgeslagen in settings.json."


st.set_page_config(page_title="TFF tariefzoeker", page_icon="EUR", layout="wide")
require_login()
st.title("TFF tariefzoeker")
st.caption("Vul eerst de levering en de zending in. De TFF/XPO-staffel wordt standaard op de achtergrond geladen.")
settings = load_settings()
saved_diesel_by_carrier = settings["diesel_by_carrier"]
saved_road_charge_by_carrier = settings["road_charge_by_carrier"]

with st.sidebar:
    st.header("Tariefbestanden")
    default_paths = [path for path in DEFAULT_WORKBOOKS if path.exists()]
    use_default = bool(default_paths)
    if use_default:
        st.success(f"{len(default_paths)} standaardstaffel(s) geladen.")
        for path in default_paths:
            st.caption(path.name)
    else:
        st.warning("Geen standaardstaffels gevonden in Downloads.")
    uploaded = st.file_uploader(
        "Optioneel: extra Excel-staffels toevoegen",
        type=["xlsx"],
        accept_multiple_files=True,
    )

countries = DEFAULT_COUNTRIES

st.subheader("1. Levering")
left, right = st.columns([1, 1])
with left:
    load_address = st.text_input("Laadadres", "Marconiweg 14, 8501 XM Joure")
    load_country = st.text_input("Laadland", "NL")
    delivery_address = st.text_input("Leveradres")
with right:
    delivery_country = st.selectbox("Leverland", countries, index=countries.index("DE") if "DE" in countries else 0)
    delivery_postcode = st.text_input("Postcode levering", "10115")
    margin_pct = st.number_input("Marge %", min_value=0.0, value=float(settings["margin_pct"]), step=0.5)

st.subheader("2. Zending")
pallet_line_count = st.number_input("Aantal verschillende palletregels", min_value=1, max_value=20, value=1, step=1)
pallet_rows: list[dict[str, Any]] = []
for idx in range(int(pallet_line_count)):
    st.markdown(f"**Palletregel {idx + 1}**")
    cols = st.columns([1, 1.2, 1, 1, 1, 1.2])
    default_type = "Europallet"
    qty = cols[0].number_input("Aantal", min_value=0, value=1 if idx == 0 else 0, step=1, key=f"qty_{idx}")
    pallet_type = cols[1].selectbox(
        "Type",
        ["Europallet", "Blokpallet", "Overig"],
        index=0 if default_type == "Europallet" else 2,
        key=f"type_{idx}",
    )
    default_length = 120 if idx == 0 else 0
    default_width = 80 if idx == 0 else 0
    length = cols[2].number_input("Lengte cm", min_value=0, value=default_length, step=1, key=f"length_{idx}")
    width = cols[3].number_input("Breedte cm", min_value=0, value=default_width, step=1, key=f"width_{idx}")
    height = cols[4].number_input("Hoogte cm", min_value=0, value=120 if idx == 0 else 0, step=1, key=f"height_{idx}")
    kg_each = cols[5].number_input("Kg per pallet", min_value=0, value=300 if idx == 0 else 0, step=1, key=f"kg_{idx}")
    pallet_rows.append(
        {
            "Aantal": qty,
            "Type": pallet_type,
            "Lengte cm": length,
            "Breedte cm": width,
            "Hoogte cm": height,
            "Kg per pallet": kg_each,
        }
    )

totals = shipment_totals(pallet_rows)
eu_pallets, block_pallets = pallet_kind_counts(pallet_rows)
manual_ldm = st.toggle("Laadmeters handmatig overschrijven", value=False)
ldm = (
    st.number_input("Laadmeters voor tariefberekening", min_value=0.0, value=round(totals["ldm"], 2), step=0.1)
    if manual_ldm
    else totals["ldm"]
)

rates: list[RateRow] = []
errors: list[str] = []
with st.spinner("Staffels laden en tarief berekenen..."):
    if uploaded:
        try:
            rates.extend(load_uploaded_rates(uploaded))
        except Exception as exc:
            errors.append(f"Upload kon niet worden gelezen: {exc}")
    if use_default:
        try:
            rates.extend(load_default_rate_books(tuple(str(path) for path in default_paths)))
        except Exception as exc:
            errors.append(f"Standaardstaffels konden niet worden gelezen: {exc}")

for error in errors:
    st.error(error)

carrier_names = sorted({row.carrier for row in rates})
diesel_by_carrier = {carrier: float(as_number(saved_diesel_by_carrier.get(carrier)) or 0.0) for carrier in carrier_names}
road_charge_by_carrier = {
    carrier: float(as_number(saved_road_charge_by_carrier.get(carrier)) or 0.0) for carrier in carrier_names
}
result_tab, diesel_tab = st.tabs(["Resultaat", "Toeslagen per vervoerder"])
with diesel_tab:
    if carrier_names:
        st.caption("Vul per vervoerder dieselpercentage en road tax in. Road tax wordt als percentage over het basistarief berekend.")
        for carrier in carrier_names:
            st.markdown(f"**{carrier}**")
            cols = st.columns(2)
            diesel_by_carrier[carrier] = cols[0].number_input(
                "Dieseltoeslag %",
                min_value=0.0,
                value=float(as_number(saved_diesel_by_carrier.get(carrier)) or 0.0),
                step=0.5,
                key=f"diesel_{carrier}",
            )
            road_charge_by_carrier[carrier] = cols[1].number_input(
                "Road tax / km-heffing %",
                min_value=0.0,
                value=float(as_number(saved_road_charge_by_carrier.get(carrier)) or 0.0),
                step=0.5,
                key=f"road_charge_{carrier}",
            )
        if st.button("Toeslagen opslaan", type="primary"):
            try:
                save_message = save_settings(diesel_by_carrier, road_charge_by_carrier, margin_pct)
                st.success(f"{save_message} Deze toeslagen worden bij de volgende refresh of herstart opnieuw geladen.")
            except requests.RequestException as exc:
                st.error(f"Opslaan in GitHub is niet gelukt: {exc}")
    else:
        st.info("Nog geen vervoerders geladen.")

zone, results = (None, [])
if rates:
    zone, results = find_rates(rates, delivery_country, delivery_postcode, totals["kg"], ldm, eu_pallets, block_pallets)
    results = sorted(
        results,
        key=lambda item: total_purchase(
            item.base_price,
            diesel_by_carrier.get(item.carrier, 0.0),
            road_charge_by_carrier.get(item.carrier, 0.0),
        ),
    )

with result_tab:
    st.divider()
    summary_cols = st.columns(6)
    summary_cols[0].metric("Land", delivery_country)
    summary_cols[1].metric("Zone", zone or "Geen match")
    summary_cols[2].metric("Pallets", f"{totals['pallets']:.0f}")
    summary_cols[3].metric("Kg", f"{totals['kg']:.0f}")
    summary_cols[4].metric("LDM", f"{ldm:.2f}")
    summary_cols[5].metric("m3", f"{totals['m3']:.2f}")

    if not rates:
        st.info("De standaardstaffel is niet gevonden. Voeg optioneel een Excel-staffel toe via de sidebar.")
        st.stop()

    if not results:
        st.warning("Geen passend tarief gevonden. Controleer land, postcode, gewicht, laadmeters en pallet-aantallen.")
        st.stop()

    best = results[0]
    best_diesel_pct = diesel_by_carrier.get(best.carrier, 0.0)
    best_road_tax_pct = road_charge_by_carrier.get(best.carrier, 0.0)
    best_road_tax_amount = road_tax_amount(best.base_price, best_road_tax_pct)
    best_purchase_total = total_purchase(best.base_price, best_diesel_pct, best_road_tax_pct)
    best_sales_total = total_with_diesel(best_purchase_total, margin_pct)
    st.success(f"Beste inkoop: {format_money(best_purchase_total)} via {best.carrier}")
    st.caption(
        "Inkoopberekening: "
        f"{format_money(best.base_price)} x (1 + {format_decimal(best_diesel_pct, 1)} / 100) "
        f"+ ({format_money(best.base_price)} x {format_decimal(best_road_tax_pct, 1)} / 100) "
        f"= {format_money(best_purchase_total)} ({format_decimal(best_purchase_total, 3)})"
    )
    st.caption(
        "Verkoopberekening: "
        f"{format_money(best_purchase_total)} x (1 + {format_decimal(margin_pct, 1)} / 100) "
        f"= {format_money(best_sales_total)} ({format_decimal(best_sales_total, 3)})"
    )

    rows_for_table = []
    for idx, row in enumerate(results, start=1):
        row_diesel_pct = diesel_by_carrier.get(row.carrier, 0.0)
        road_tax_pct = road_charge_by_carrier.get(row.carrier, 0.0)
        road_tax = road_tax_amount(row.base_price, road_tax_pct)
        purchase_total = total_purchase(row.base_price, row_diesel_pct, road_tax_pct)
        sales_total = total_with_diesel(purchase_total, margin_pct)
        rows_for_table.append(
            {
                "Rank": idx,
                "Vervoerder": row.carrier,
                "Basistarief": format_money(row.base_price),
                "Diesel %": f"{row_diesel_pct:.1f}%",
                "Diesel bedrag": format_money(total_with_diesel(row.base_price, row_diesel_pct) - row.base_price),
                "Road tax %": f"{road_tax_pct:.1f}%",
                "Road tax bedrag": format_money(road_tax),
                "Inkoop incl. diesel": format_money(purchase_total),
                "Marge %": f"{margin_pct:.1f}%",
                "Verkoop": format_money(sales_total),
                "Land": row.country,
                "Zone": row.zone,
                "Staffel kg": row.max_kg,
                "Staffel ldm": row.max_ldm,
                "Europallets t/m": row.max_eu_pallets,
                "Blokpallets t/m": row.max_block_pallets,
                "Transit": row.transit or "",
                "Bron": row.source,
            }
        )

    st.dataframe(rows_for_table, width="stretch", hide_index=True)

    with st.expander("Berekening"):
        st.write(
            {
                "laadadres": load_address,
                "laadland": load_country.upper(),
                "leveradres": delivery_address,
                "leverpostcode": delivery_postcode,
                "postcode_match": zone,
                "brutogewicht_kg": round(totals["kg"], 3),
                "volume_m3": round(totals["m3"], 3),
                "laadmeters": round(ldm, 3),
                "europallets": eu_pallets,
                "blokpallets": block_pallets,
                "dieseltoeslag_pct": best_diesel_pct,
                "road_tax_pct": best_road_tax_pct,
                "road_tax_bedrag": round(best_road_tax_amount, 3),
                "marge_pct": margin_pct,
                "formule_inkoop": "basistarief * (1 + dieseltoeslag_pct / 100) + basistarief * (road_tax_pct / 100)",
                "formule_verkoop": "inkoop_incl_diesel * (1 + marge_pct / 100)",
                "ingevulde_inkoopberekening": (
                    f"{best.base_price} * (1 + {best_diesel_pct} / 100) + {best.base_price} * ({best_road_tax_pct} / 100) = {round(best_purchase_total, 3)}"
                ),
                "ingevulde_verkoopberekening": (
                    f"{round(best_purchase_total, 3)} * (1 + {margin_pct} / 100) = {round(best_sales_total, 3)}"
                ),
            }
        )
